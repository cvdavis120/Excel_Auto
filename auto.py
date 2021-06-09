# Munipulating XCEL WITH PYTHON
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook
# make a workbook
# wb = Workbook()
# make a worksheet
#load existing spreadsheet
# use absolute path name
wb=load_workbook('test.xlsx')
ws = wb.active
# GET THE SEARCH TERMS IN A COLUMN
new_arr = ws.iter_cols(2, 2,2,False)

# WORD BANKS
ad_group_wb = ['milwaukee', 'dewalt', 'ingersoll', 'rand', 'mac']
aircat_wb = ['aircat', "cat"]
primary_cordless_wb=['cordless','electric','battery']


# AD GROUP SEARCH
def ad_group_sort(search, my_row):
    words = search.split()
    for i in words:
        if (i == "electric" or "elec"):
            ws.cell(row=my_row, column=5).value = "ELECTRIC"
    for i in words:
        if (i in ad_group_wb):
            ws.cell(row=my_row, column=3).value = 'COMPETITOR'
            ws.cell(row=my_row, column=4).value = i.capitalize()
            return
        elif (i in aircat_wb):
            ws.cell(row=my_row, column=3).value = 'AIRCAT'
            ws.cell(row=my_row, column=4).value = 'AIRCAT'
            return
        else:
            ws.cell(row=my_row, column=3).value = 'GENERIC'
            #CORDLESS PRIMARY
            primary_cordless_sort(i,my_row)

    
    return

def primary_cordless_sort(search,my_row):
    if (search in primary_cordless_wb):
        ws.cell(row=my_row, column=4).value = "CORDLESS"
    return
   



for i in (next(new_arr)):
    ad_group_sort(i.value,i.row)

        

wb.save('test_update.xlsx')